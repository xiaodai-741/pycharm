import os
import re

import pandas as pd
import pymysql as pymysql
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

direName = '新客户签约2'
# 创建excel文件
df = pd.DataFrame()
df.to_excel(f'./{direName}.xlsx')
wb = load_workbook(rf'./{direName}.xlsx')
sheet = wb.active
indexList = ['id', '客户名', '区域', '业务员', '任务', '首款', '客户等级']

for i in range(1, len(indexList)):
    sheet[f'{get_column_letter(i)}1'] = indexList[i - 1]
active_number = sheet.max_row + 1

# 读取数据
for file in os.listdir(direName):
    print(file)
    wwb = load_workbook(f'{direName}/{file}', data_only=True)
    wwb_sheet = wwb.active
    print(re.findall(r"业务人员： *(.+) *", wwb_sheet['a2'].value)[0].strip())
    sheet[f'A{active_number}'] = active_number - 1
    sheet[f'b{active_number}'] = wwb_sheet['c3'].value
    sheet[f'c{active_number}'] = re.findall(r"大区：(.+?)业务", wwb_sheet['a2'].value)[0].strip()
    sheet[f'd{active_number}'] = re.findall(r"业务人员： *(.+) *", wwb_sheet['a2'].value)[0].strip()
    sheet[f'e{active_number}'] = wwb_sheet['c11'].value
    sheet[f'f{active_number}'] = wwb_sheet['e11'].value
    sheet[f'g{active_number}'] = wwb_sheet['c10'].value
    active_number += 1
wb.save(f'./{direName}.xlsx')
