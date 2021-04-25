import openpyxl
import xlwt
import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('客户区域划分.py')
sheet = wb.active
