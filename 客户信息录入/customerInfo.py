import os
import re

import pandas as pd
import pymysql as pymysql
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

direName = '客户信息录入'
# 创建excel文件
df = pd.DataFrame()
df.to_excel(f'./{direName}.xlsx')
wb = load_workbook(rf'./{direName}.xlsx')
sheet = wb.active
indexList = ['id', 'customer', 'linkman', 'phone_number', 'number_of_employee', 'company_address',
             'real_boss_and_phone', 'boss_birthday', 'boss_chinese_zodiac', 'company_birthday',
             'area', 'saleman', 'sale_product_all_is_qb', 'other_product_name', 'total_annual_sales',
             'Products_accounted_for_qb', 'Products_accounted_for_chenpi', 'Products_accounted_for_ganpucha',
             'area_responsible', 'customer_level', 'task_money', 'finish_money', 'purchase_amount_of_qiyueguo',
             'the_storage_amount_of_xinpi', 'the_qbproduct_sales_first', 'the_qbproduct_sales_second',
             'the_qbproduct_sales_third', 'fenxiao_number_name', 'zhuangui_number_name',
             'customer_specific_circumstance', 'question', 'remark']

for i in range(1, len(indexList)):
    sheet[f'{get_column_letter(i)}1'] = indexList[i - 1]
active_number = sheet.max_row + 1

# 读取数据
# for file in os.listdir(direName):
#     wwb = load_workbook(f'{direName}/{file}', data_only=True)
#     wwb_sheet = wwb.active
wwb = load_workbook('客户详细信息资料统计表 - 王立城(2)(1).xlsx', data_only=True)
wwb_sheet = wwb.active
# print(wwb_sheet['a2'].value)
# b = re.findall(r"大区：(.+?)业务", wwb_sheet['a2'].value)
# c = b[0]
# print(c.strip())
# a = '123a231w'
# print(re.findall('a(.+?)w', a))
a = wwb_sheet['a2'].value
# print(re.findall(r"业务人员：(.{10})", wwb_sheet['a2'].value)[0].strip())
b = wwb_sheet['d13'].value
c = wwb_sheet['d14'].value
d = wwb_sheet['d15'].value
e = "1."+b+"\n"+"2."+c+"\n"+"3."+d
# print(e)
sheet[f'A{active_number }'] = active_number-1
sheet[f'b{active_number }'] = wwb_sheet['c3'].value
sheet[f'c{active_number }'] = wwb_sheet['e3'].value
sheet[f'd{active_number }'] = wwb_sheet['g3'].value
sheet[f'e{active_number }'] = wwb_sheet['e4'].value
sheet[f'f{active_number }'] = wwb_sheet['c4'].value
sheet[f'g{active_number }'] = wwb_sheet['g4'].value
sheet[f'h{active_number }'] = wwb_sheet['c5'].value
sheet[f'i{active_number }'] = wwb_sheet['e5'].value
sheet[f'j{active_number }'] = wwb_sheet['g5'].value
sheet[f'k{active_number }'] = re.findall(r"大区：(.+?)业务", wwb_sheet['a2'].value)[0].strip()
sheet[f'l{active_number }'] = a[a.find('业务人员：')+5:]
sheet[f'm{active_number }'] = wwb_sheet['c6'].value
sheet[f'n{active_number }'] = wwb_sheet['e6'].value
sheet[f'o{active_number }'] = wwb_sheet['g6'].value
sheet[f'p{active_number }'] = wwb_sheet['c7'].value
sheet[f'q{active_number }'] = wwb_sheet['e7'].value
sheet[f'r{active_number }'] = wwb_sheet['g7'].value
sheet[f's{active_number }'] = wwb_sheet['c8'].value
sheet[f't{active_number }'] = wwb_sheet['e8'].value
sheet[f'u{active_number }'] = wwb_sheet['g8'].value
sheet[f'v{active_number }'] = wwb_sheet['c9'].value
sheet[f'w{active_number }'] = wwb_sheet['e9'].value
sheet[f'x{active_number }'] = wwb_sheet['g9'].value
sheet[f'y{active_number }'] = wwb_sheet['c10'].value
sheet[f'z{active_number }'] = wwb_sheet['e10'].value
sheet[f'aa{active_number }'] = wwb_sheet['g10'].value

sheet[f'ab{active_number }'] = wwb_sheet['d11'].value
sheet[f'ac{active_number }'] = wwb_sheet['d12'].value

sheet[f'ad{active_number }'] = e
sheet[f'ae{active_number }'] = wwb_sheet['c16'].value
sheet[f'af{active_number }'] = wwb_sheet['b17'].value


active_number += 1
#
wb.save(f'./{direName}.xlsx')
# df = pd.read_excel('./2021年1月陈皮酒订单.xlsx')
# df.to_sql(name='chenpijiudingdan', con='mysql+pymysql://root:123456@localhost:3306/ChenPiJiu?charset=utf8',
#           if_exists='replace',
#           index=False)
