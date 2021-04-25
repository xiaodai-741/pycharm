import os
import pandas as pd
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

indexList = ['id', 'customer', 'linkman', 'phone_number', 'number_of_employee', 'company_address',
             'real_boss_and_phone', 'boss_birthday', 'boss_chinese_zodiac', 'company_birthday',
             'area', 'saleman', 'sale_product_all_is_qb', 'other_product_name', 'total_annual_sales',
             'Products_accounted_for_qb', 'Products_accounted_for_chenpi', 'Products_accounted_for_ganpucha',
             'area_responsible', 'customer_level', 'task_money', 'finish_money', 'purchase_amount_of_qiyueguo',
             'the_storage_amount_of_xinpi', 'the_qbproduct_sales_first', 'the_qbproduct_sales_second',
             'the_qbproduct_sales_third', 'fenxiao_number_name', 'zhuangui_number_name',
             'customer_specific_circumstance', 'question', 'remark']
df = pd.DataFrame()
df.to_excel(f'./客户信息总表.xlsx')
customer_wb = load_workbook(rf'./客户信息总表.xlsx')
customer_sheet = customer_wb.active

for i in range(1, len(indexList)):
    customer_sheet[f'{get_column_letter(i)}1'] = indexList[i - 1]
customer_number = customer_sheet.max_row + 1
for area in os.listdir('./各区客户资料表'):
    print('\n' + area)
    df = pd.DataFrame()
    df.to_excel(f'./{area}.xlsx')
    wb = load_workbook(rf'./{area}.xlsx')
    sheet = wb.active

    for i in range(1, len(indexList)):
        sheet[f'{get_column_letter(i)}1'] = indexList[i - 1]
        active_number = sheet.max_row + 1
    for customer_info_list in os.listdir(f'./各区客户资料表/{area}'):
        print(customer_info_list)
        wbb = openpyxl.load_workbook(f'./各区客户资料表/{area}/{customer_info_list}')
        sheets = wbb.sheetnames
        for i in range(len(sheets)):
            wwb_sheet = wbb[sheets[i]]
            a = str(wwb_sheet['a2'].value)
            # print(re.findall(r"业务人员：(.{10})", wwb_sheet['a2'].value)[0].strip())
            b = wwb_sheet['d13'].value
            c = wwb_sheet['d14'].value
            d = wwb_sheet['d15'].value
            e = "1." + str(b) + "\n" + "2." + str(c) + "\n" + "3." + str(d)
            # print(e)
            sheet[f'A{active_number}'] = active_number - 1
            sheet[f'b{active_number}'] = wwb_sheet['c3'].value
            sheet[f'c{active_number}'] = wwb_sheet['e3'].value
            sheet[f'd{active_number}'] = wwb_sheet['g3'].value
            sheet[f'e{active_number}'] = wwb_sheet['e4'].value
            sheet[f'f{active_number}'] = wwb_sheet['c4'].value
            sheet[f'g{active_number}'] = wwb_sheet['g4'].value
            sheet[f'h{active_number}'] = wwb_sheet['c5'].value
            sheet[f'i{active_number}'] = wwb_sheet['e5'].value
            sheet[f'j{active_number}'] = wwb_sheet['g5'].value
            # re.findall(r"大区：(.+?)业务", wwb_sheet['a2'].value)[0].strip()
            sheet[f'k{active_number}'] = area
            sheet[f'l{active_number}'] = a[a.find('业务人员：') + 5:]
            sheet[f'm{active_number}'] = wwb_sheet['c6'].value
            sheet[f'n{active_number}'] = wwb_sheet['e6'].value
            sheet[f'o{active_number}'] = wwb_sheet['g6'].value
            sheet[f'p{active_number}'] = wwb_sheet['c7'].value
            sheet[f'q{active_number}'] = wwb_sheet['e7'].value
            sheet[f'r{active_number}'] = wwb_sheet['g7'].value
            sheet[f's{active_number}'] = wwb_sheet['c8'].value
            sheet[f't{active_number}'] = wwb_sheet['e8'].value
            sheet[f'u{active_number}'] = wwb_sheet['g8'].value
            sheet[f'v{active_number}'] = wwb_sheet['c9'].value
            sheet[f'w{active_number}'] = wwb_sheet['e9'].value
            sheet[f'x{active_number}'] = wwb_sheet['g9'].value
            sheet[f'y{active_number}'] = wwb_sheet['c10'].value
            sheet[f'z{active_number}'] = wwb_sheet['e10'].value
            sheet[f'aa{active_number}'] = wwb_sheet['g10'].value
            sheet[f'ab{active_number}'] = wwb_sheet['d11'].value
            sheet[f'ac{active_number}'] = wwb_sheet['d12'].value
            sheet[f'ad{active_number}'] = e
            sheet[f'ae{active_number}'] = wwb_sheet['c16'].value
            sheet[f'af{active_number}'] = wwb_sheet['b17'].value

            customer_sheet[f'A{customer_number}'] = customer_number - 1
            customer_sheet[f'b{customer_number}'] = wwb_sheet['c3'].value
            customer_sheet[f'c{customer_number}'] = wwb_sheet['e3'].value
            customer_sheet[f'd{customer_number}'] = wwb_sheet['g3'].value
            customer_sheet[f'e{customer_number}'] = wwb_sheet['e4'].value
            customer_sheet[f'f{customer_number}'] = wwb_sheet['c4'].value
            customer_sheet[f'g{customer_number}'] = wwb_sheet['g4'].value
            customer_sheet[f'h{customer_number}'] = wwb_sheet['c5'].value
            customer_sheet[f'i{customer_number}'] = wwb_sheet['e5'].value
            customer_sheet[f'j{customer_number}'] = wwb_sheet['g5'].value
            # re.findall(r"大区：(.+?)业务", wwb_sheet['a2'].value)[0].strip()
            customer_sheet[f'k{customer_number}'] = area
            customer_sheet[f'l{customer_number}'] = a[a.find('业务人员：') + 5:]
            customer_sheet[f'm{customer_number}'] = wwb_sheet['c6'].value
            customer_sheet[f'n{customer_number}'] = wwb_sheet['e6'].value
            customer_sheet[f'o{customer_number}'] = wwb_sheet['g6'].value
            customer_sheet[f'p{customer_number}'] = wwb_sheet['c7'].value
            customer_sheet[f'q{customer_number}'] = wwb_sheet['e7'].value
            customer_sheet[f'r{customer_number}'] = wwb_sheet['g7'].value
            customer_sheet[f's{customer_number}'] = wwb_sheet['c8'].value
            customer_sheet[f't{customer_number}'] = wwb_sheet['e8'].value
            customer_sheet[f'u{customer_number}'] = wwb_sheet['g8'].value
            customer_sheet[f'v{customer_number}'] = wwb_sheet['c9'].value
            customer_sheet[f'w{customer_number}'] = wwb_sheet['e9'].value
            customer_sheet[f'x{customer_number}'] = wwb_sheet['g9'].value
            customer_sheet[f'y{customer_number}'] = wwb_sheet['c10'].value
            customer_sheet[f'z{customer_number}'] = wwb_sheet['e10'].value
            customer_sheet[f'aa{customer_number}'] = wwb_sheet['g10'].value
            customer_sheet[f'ab{customer_number}'] = wwb_sheet['d11'].value
            customer_sheet[f'ac{customer_number}'] = wwb_sheet['d12'].value
            customer_sheet[f'ad{customer_number}'] = e
            customer_sheet[f'ae{customer_number}'] = wwb_sheet['c16'].value
            customer_sheet[f'af{customer_number}'] = wwb_sheet['b17'].value

            active_number += 1
            customer_number += 1
        wb.save(f'./{area}.xlsx')
customer_wb.save(f'./客户信息总表.xlsx')
# df = pd.read_excel('./2021年1月陈皮酒订单.xlsx')
# df.to_sql(name='chenpijiudingdan', con='mysql+pymysql://root:123456@localhost:3306/ChenPiJiu?charset=utf8',
#           if_exists='replace',index=False)