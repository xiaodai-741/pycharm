import re
from openpyxl import load_workbook

wbb =  load_workbook('./各区客户资料表/华东/袁梦洁客户总结.xlsx')
sheets = wbb.sheetnames

# wwb_sheet = wbb.active
# a = re.findall(r"大区：(.+?)业务", wwb_sheet['a2'].value)[0].strip()


print(sheets)

