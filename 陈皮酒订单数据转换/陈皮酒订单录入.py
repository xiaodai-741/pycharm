import os
import pandas as pd
import pymysql as pymysql
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

direName = '2021年4月陈皮酒订单'
# # 创建excel文件
# df = pd.DataFrame()
# df.to_excel(f'./{direName}.xlsx')
wb = load_workbook(rf'./{direName}.xlsx')
sheet = wb.active
# indexList = ['日期', '区域', '城市经理', '大区经理', '订货单位', '产品名称', '数量（瓶）', '单价', '金额', '搭赠数量（瓶）',
#              '搭赠单价', '搭赠金额', '备注', '收货地址', '收货人/电话']

# for i in range(1, 16):
#     sheet[f'{get_column_letter(i)}1'] = indexList[i - 1]
active_number = sheet.max_row

# 读取数据
date_name = r'E:\pycharm\陈皮酒订单数据转换\2021年四月\2021年第三周订单'
for file in os.listdir(date_name):
    wwb = load_workbook(f'{date_name}/{file}', data_only=True)
    wwb_sheet = wwb.active
    # wwb = load_workbook('2021年1月陈皮酒订单/2021.1.7陈皮酒订单    山西运福源商贸有限公司.xlsx', data_only=True)
    # wwb_sheet = wwb.active
    for i in range(4, 12):
        if wwb_sheet[f'b{i}'].value is None and wwb_sheet[f'b{i + 1}'].value is None:
            break
        else:
            sheet[f'A{active_number + 1}'] = wwb_sheet['k2'].value
            sheet[f'b{active_number + 1}'] = wwb_sheet['b2'].value
            sheet[f'c{active_number + 1}'] = wwb_sheet['e2'].value
            sheet[f'd{active_number + 1}'] = wwb_sheet['i2'].value
            sheet[f'e{active_number + 1}'] = wwb_sheet['c36'].value
            sheet[f'f{active_number + 1}'] = wwb_sheet[f'b{i}'].value
            sheet[f'g{active_number + 1}'] = wwb_sheet[f'h{i}'].value
            sheet[f'h{active_number + 1}'] = wwb_sheet[f'g{i}'].value
            sheet[f'i{active_number + 1}'] = wwb_sheet[f'i{i}'].value
            sheet[f'j{active_number + 1}'] = 0
            sheet[f'k{active_number + 1}'] = 0
            sheet[f'l{active_number + 1}'] = 0
            sheet[f'm{active_number + 1}'] = wwb_sheet[f'k{i}'].value
            sheet[f'n{active_number + 1}'] = wwb_sheet['c35'].value
            sheet[f'o{active_number + 1}'] = wwb_sheet['c34'].value
            active_number += 1
    for i in range(12, 15):
        if wwb_sheet[f'b{i}'].value is None and wwb_sheet[f'b{i + 1}'].value is None:
            break
        else:
            sheet[f'A{active_number + 1}'] = wwb_sheet['k2'].value
            sheet[f'b{active_number + 1}'] = wwb_sheet['b2'].value
            sheet[f'c{active_number + 1}'] = wwb_sheet['e2'].value
            sheet[f'd{active_number + 1}'] = wwb_sheet['i2'].value
            sheet[f'e{active_number + 1}'] = wwb_sheet['c36'].value
            sheet[f'f{active_number + 1}'] = wwb_sheet[f'b{i}'].value
            sheet[f'g{active_number + 1}'] = wwb_sheet[f'h{i}'].value
            sheet[f'h{active_number + 1}'] = wwb_sheet[f'g{i}'].value
            sheet[f'i{active_number + 1}'] = wwb_sheet[f'i{i}'].value
            sheet[f'j{active_number + 1}'] = 0
            sheet[f'k{active_number + 1}'] = 0
            sheet[f'l{active_number + 1}'] = 0
            sheet[f'm{active_number + 1}'] = wwb_sheet[f'k{i}'].value
            sheet[f'n{active_number + 1}'] = wwb_sheet['c35'].value
            sheet[f'o{active_number + 1}'] = wwb_sheet['c34'].value
            active_number += 1
    for i in range(18, 25):
        if wwb_sheet[f'b{i}'].value is None and wwb_sheet[f'b{i + 1}'].value is None:
            break
        else:
            sheet[f'A{active_number + 1}'] = wwb_sheet['k2'].value
            sheet[f'b{active_number + 1}'] = wwb_sheet['b2'].value
            sheet[f'c{active_number + 1}'] = wwb_sheet['e2'].value
            sheet[f'd{active_number + 1}'] = wwb_sheet['i2'].value
            sheet[f'e{active_number + 1}'] = wwb_sheet['c36'].value
            sheet[f'f{active_number + 1}'] = wwb_sheet[f'b{i}'].value
            sheet[f'g{active_number + 1}'] = 0
            sheet[f'h{active_number + 1}'] = 0
            sheet[f'i{active_number + 1}'] = 0
            sheet[f'j{active_number + 1}'] = wwb_sheet[f'h{i}'].value
            sheet[f'k{active_number + 1}'] = wwb_sheet[f'g{i}'].value
            sheet[f'l{active_number + 1}'] = wwb_sheet[f'i{i}'].value
            sheet[f'm{active_number + 1}'] = wwb_sheet[f'k{i}'].value
            sheet[f'n{active_number + 1}'] = wwb_sheet['c35'].value
            sheet[f'o{active_number + 1}'] = wwb_sheet['c34'].value
            active_number += 1
    for i in range(25, 28):
        if wwb_sheet[f'b{i}'].value is None and wwb_sheet[f'b{i + 1}'].value is None:
            break
        else:
            sheet[f'A{active_number + 1}'] = wwb_sheet['k2'].value
            sheet[f'b{active_number + 1}'] = wwb_sheet['b2'].value
            sheet[f'c{active_number + 1}'] = wwb_sheet['e2'].value
            sheet[f'd{active_number + 1}'] = wwb_sheet['i2'].value
            sheet[f'e{active_number + 1}'] = wwb_sheet['c36'].value
            sheet[f'f{active_number + 1}'] = wwb_sheet[f'b{i}'].value
            sheet[f'g{active_number + 1}'] = 0
            sheet[f'h{active_number + 1}'] = 0
            sheet[f'i{active_number + 1}'] = 0
            sheet[f'j{active_number + 1}'] = wwb_sheet[f'h{i}'].value
            sheet[f'k{active_number + 1}'] = wwb_sheet[f'g{i}'].value
            sheet[f'l{active_number + 1}'] = wwb_sheet[f'i{i}'].value
            sheet[f'm{active_number + 1}'] = wwb_sheet[f'k{i}'].value
            sheet[f'n{active_number + 1}'] = wwb_sheet['c35'].value
            sheet[f'o{active_number + 1}'] = wwb_sheet['c34'].value
            active_number += 1
wb.save(f'./{direName}.xlsx')

