import os
import re

import pandas as pd
import pymysql as pymysql
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

direName = '库存信息盘点表'
# 创建excel文件
df = pd.DataFrame()
df.to_excel(f'./{direName}.xlsx')
wb = load_workbook(rf'./{direName}.xlsx')
sheet = wb.active
indexList = ['序号', '经销商名称', '负责业务', '产品名称', '当前产品数量', '库存占比', '是否为囤货', '备注']

for i in range(1, len(indexList) + 1):
    sheet[f'{get_column_letter(i)}1'] = indexList[i - 1]
active_number = sheet.max_row + 1

# 读取数据
file_base = '盘点表华东'
for file in os.listdir(file_base):
    print(file.title())
    wwb = load_workbook(f'{file_base}/{file}', data_only=True)
    for i in wwb:
        wwb_sheet = wwb[i.title]
        print(i.title)
        customer = re.findall(r"经销商名称：(.+?)负责业务", wwb_sheet['a2'].value)[0].strip()
        saleman = re.findall(r"负责业务：(.+?)盘点时间", wwb_sheet['a2'].value)[0].strip()
        for j in range(4,wwb_sheet.max_row):
            product = wwb_sheet[f'b{j}'].value
            number = wwb_sheet[f'c{j}'].value
            percentage = wwb_sheet[f'D{j}'].value
            tunhuo = wwb_sheet[f'E{j}'].value
            remake = wwb_sheet[f'F{j}'].value
            if product is not None:
                print(product)
                print(customer)
                sheet[f'A{active_number}'] = active_number - 1
                sheet[f'B{active_number}'] = customer
                sheet[f'C{active_number}'] = saleman
                sheet[f'D{active_number}'] = product
                sheet[f'E{active_number}'] = number
                sheet[f'F{active_number}'] = percentage
                sheet[f'G{active_number}'] = tunhuo
                sheet[f'H{active_number}'] = remake
                active_number += 1
            else:
                break
wb.save(f'./{direName}.xlsx')
