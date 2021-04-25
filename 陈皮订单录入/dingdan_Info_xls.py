import os
import re

import pandas as pd
import pymysql as pymysql
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

direName = '3.14陈皮订单1'
# 创建excel文件
df = pd.DataFrame()
df.to_excel(f'./{direName}.xlsx')
wb = load_workbook(rf'./{direName}.xlsx')
sheet = wb.active
indexList = ['id', '订货单位', '区域', '业务员', '产品名称', '规格', '单位', '出厂价', '数量', '金额（元）', '订单日期', '备注']

for i in range(1, len(indexList)):
    sheet[f'{get_column_letter(i)}1'] = indexList[i - 1]
active_number = sheet.max_row + 1
file_name = r"E:\pycharm\陈皮订单录入\3月订单\3.14订单1"
# 读取数据
for file in os.listdir(file_name):
    print(file)
    wwb = load_workbook(rf"E:\pycharm\陈皮订单录入\3月订单\3.14订单1\{file}", data_only=True)
    wwb_sheet = wwb.active
    customer = re.findall(r"陈皮订单 *(.+?).xlsx", file)[0].strip()
    area = re.findall(r"区域：(.+?)业务", wwb_sheet['a2'].value)[0].strip()
    saleman = re.findall(r"业务人员：(.+?)所属", wwb_sheet['a2'].value)[0].strip()
    date = re.findall(r"订单日期： *(.+) *", wwb_sheet['a2'].value)[0].strip()
    for i in range(4, 50):
        if wwb_sheet[f'b{i}'].value is None and i != 4:
            break
        else:
            sheet[f'A{active_number}'] = active_number - 1
            sheet[f'b{active_number}'] = customer
            sheet[f'c{active_number}'] = area
            sheet[f'd{active_number}'] = saleman
            sheet[f'e{active_number}'] = wwb_sheet[f'b{i}'].value
            sheet[f'f{active_number}'] = wwb_sheet[f'c{i}'].value
            sheet[f'g{active_number}'] = wwb_sheet[f'd{i}'].value
            sheet[f'h{active_number}'] = wwb_sheet[f'e{i}'].value
            sheet[f'i{active_number}'] = wwb_sheet[f'f{i}'].value
            sheet[f'j{active_number}'] = wwb_sheet[f'g{i}'].value
            sheet[f'k{active_number}'] = date
            sheet[f'l{active_number}'] = wwb_sheet[f'h{i}'].value
            active_number += 1
wb.save(f'./{direName}.xlsx')
